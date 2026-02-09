import streamlit as st
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import re
from datetime import datetime
import traceback
import zipfile
from io import BytesIO

# --- Función para Normalizar Claves ---
def normalize_key(text):
    if not isinstance(text, str):
        text = str(text)
    text = text.lower()
    replacements = {
        " ": "_", "º": "nro", ".": "", ":": "", "ñ": "n", "ó": "o", "ö": "o",
        "é": "e", "í": "i", "á": "a", "ú": "u", "ü": "u", "-": "_", "(": "", ")": "",
        "/": "_"
    }
    for char, replacement in replacements.items():
        text = text.replace(char, replacement)
    text = re.sub(r'_+', '_', text)
    return text.strip("_")

# --- Función para Convertir Letra de Columna a Índice ---
def col_letter_to_index(col_letter):
    """Convierte letra de columna Excel (ej: 'A', 'AB') a índice numérico (1-based)"""
    index = 0
    for i, char in enumerate(reversed(col_letter.upper())):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index

# --- Función para Validar Plan de Acción ---
def validar_plan_accion_general(get_cell_func, sheet_name, fila_idx, columnas_obligatorias):
    """
    Verifica si existen contenidos en las columnas obligatorias para un plan de acción.
    Retorna: 'completo', 'incompleto', 'sin_plan'
    """
    columnas_con_contenido = 0
    total_columnas = len(columnas_obligatorias)
    
    for col_letter in columnas_obligatorias:
        col_idx = col_letter_to_index(col_letter)
        valor_str = get_cell_func(sheet_name, fila_idx, col_idx)
        if valor_str:  # Si tiene contenido válido
            columnas_con_contenido += 1
    
    if columnas_con_contenido == total_columnas:
        return 'completo'
    elif columnas_con_contenido > 0:
        return 'incompleto'
    else:
        return 'sin_plan'

# --- Función para Validar Identificación Inicial (Hoja 3) ---
def validar_identificacion_inicial(get_cell_func, sheet_exists_func):
    """
    Valida la Hoja "3" (Identificación Inicial).
    
    Regla: Si las celdas C y D de una fila tienen datos (distintos de "0" y no vacías),
    entonces las columnas E hasta K de esa fila deben contener "SI" o "NO".
    
    Retorna: Lista de diccionarios con casos que tienen problemas
    """
    alertas = []
    
    # Verificar si existe la hoja "3"
    if not sheet_exists_func("3"):
        return alertas  # Retornar lista vacía si no existe la hoja
    
    # Columnas a verificar: B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11
    COL_CASO = 2      # B - Número de caso
    COL_C = 3         # C
    COL_D = 4         # D
    COLS_EVALUAR = [5, 6, 7, 8, 9, 10, 11]  # E, F, G, H, I, J, K
    
    VALORES_VALIDOS = ["SI", "NO", "SÍ"]  # Valores aceptados (incluyendo "SÍ" con acento)
    
    # Recorrer filas 14 a 3013
    for fila in range(14, 3014):
        # Obtener valores de columnas C y D
        valor_c = get_cell_func("3", fila, COL_C)
        valor_d = get_cell_func("3", fila, COL_D)
        
        # Verificar si C o D tienen datos (distintos de "0" y no vacías)
        tiene_datos_c = valor_c and valor_c != "0"
        tiene_datos_d = valor_d and valor_d != "0"
        
        if tiene_datos_c or tiene_datos_d:
            # Obtener número de caso
            num_caso = get_cell_func("3", fila, COL_CASO)
            if not num_caso:
                num_caso = f"Fila {fila}"
            
            # Verificar que columnas E-K tengan "SI" o "NO"
            columnas_incompletas = []
            for col_idx in COLS_EVALUAR:
                valor_col = get_cell_func("3", fila, col_idx).upper() if get_cell_func("3", fila, col_idx) else ""
                if valor_col not in VALORES_VALIDOS:
                    # Convertir índice a letra de columna
                    col_letra = chr(ord('A') + col_idx - 1)
                    columnas_incompletas.append(col_letra)
            
            # Si hay columnas sin completar, agregar alerta
            if columnas_incompletas:
                alertas.append({
                    "caso": num_caso,
                    "fila": fila,
                    "columnas_faltantes": columnas_incompletas,
                    "mensaje": f"Caso {num_caso}: Falta completar identificación inicial"
                })
    
    return alertas

# --- Función para Validar Cruzada Hoja 3 vs Hojas de Factores ---
def validar_evaluaciones_pendientes(get_cell_func, sheet_exists_func):
    """
    Valida que si en Hoja 3 se marcó "SI" para un factor de riesgo,
    entonces debe existir una evaluación en la hoja correspondiente.
    
    Mapeo:
    - Columna E (Repetitividad) → Hoja 4
    - Columna F (Postura) → Hoja 5
    - Columna G (MMC LDT) → Hoja 6
    - Columna H (MMC EA) → Hoja 7
    - Columna I (MMP) → Hoja 8
    - Columna J (Vibración CC) → Hoja 10
    - Columna K (Vibración MB) → Hoja 9
    
    Retorna: Lista de diccionarios con evaluaciones pendientes
    """
    alertas = []
    
    # Verificar si existe la hoja "3"
    if not sheet_exists_func("3"):
        return alertas
    
    # Mapeo de columnas de Hoja 3 a hojas de factores
    MAPEO_FACTORES = {
        5: {"factor": "Repetitividad", "hoja": "4", "rango_filas": (14, 116)},      # E
        6: {"factor": "Postura", "hoja": "5", "rango_filas": (17, 116)},            # F
        7: {"factor": "MMC LDT", "hoja": "6", "rango_filas": (18, 118)},            # G
        8: {"factor": "MMC EA", "hoja": "7", "rango_filas": (17, 117)},             # H
        9: {"factor": "MMP", "hoja": "8", "rango_filas": (17, 117)},                # I
        10: {"factor": "Vibración CC", "hoja": "10", "rango_filas": (16, 116)},     # J
        11: {"factor": "Vibración MB", "hoja": "9", "rango_filas": (16, 116)}       # K
    }
    
    COL_CASO_H3 = 2  # Columna B en Hoja 3
    COL_NRO_FACTOR = 2  # Columna B en hojas de factores
    
    # Construir diccionario de casos evaluados por cada hoja de factor
    casos_evaluados_por_hoja = {}
    for col_idx, config in MAPEO_FACTORES.items():
        hoja_factor = config["hoja"]
        if not sheet_exists_func(hoja_factor):
            casos_evaluados_por_hoja[hoja_factor] = set()
            continue
        
        casos_en_hoja = set()
        rango_inicio, rango_fin = config["rango_filas"]
        for fila in range(rango_inicio, rango_fin):
            nro_caso = get_cell_func(hoja_factor, fila, COL_NRO_FACTOR)
            if nro_caso and nro_caso != "0":
                casos_en_hoja.add(str(nro_caso).strip())
        casos_evaluados_por_hoja[hoja_factor] = casos_en_hoja
    
    # Recorrer Hoja 3 y verificar evaluaciones pendientes
    for fila in range(14, 3014):
        # Obtener número de caso
        num_caso = get_cell_func("3", fila, COL_CASO_H3)
        if not num_caso or num_caso == "0":
            continue
        
        num_caso_str = str(num_caso).strip()
        
        # Verificar cada factor
        for col_idx, config in MAPEO_FACTORES.items():
            valor_identificacion = get_cell_func("3", fila, col_idx)
            valor_upper = valor_identificacion.upper() if valor_identificacion else ""
            
            # Si está marcado como "SI", verificar que exista evaluación
            if valor_upper in ["SI", "SÍ"]:
                hoja_factor = config["hoja"]
                casos_evaluados = casos_evaluados_por_hoja.get(hoja_factor, set())
                
                if num_caso_str not in casos_evaluados:
                    alertas.append({
                        "caso": num_caso_str,
                        "fila": fila,
                        "factor": config["factor"],
                        "hoja": hoja_factor,
                        "mensaje": f"Caso {num_caso_str}: Marcado SI en {config['factor']} pero sin evaluación en Hoja {hoja_factor}"
                    })
    
    return alertas

# --- Función para Procesar Excel y Extraer Datos ---
def procesar_excel_resumen(uploaded_excel_file):
    """
    Procesa el Excel y extrae:
    1. Información general de la empresa
    2. Niveles de riesgo por puesto
    3. Validación de planes de acción para riesgos críticos
    """
    if uploaded_excel_file is None:
        return None
    
    # Guardar el archivo en memoria para múltiples intentos
    try:
        file_content = uploaded_excel_file.read()
        uploaded_excel_file.seek(0)
        
        wb = None
        hojas_pandas = {}
        usar_pandas = False
        error_messages = []
        
        # INTENTO 1: calamine (motor más robusto)
        try:
            st.info("🔄 Cargando archivo con motor calamine...")
            hojas_pandas = pd.read_excel(
                BytesIO(file_content),
                sheet_name=None,  # Leer todas las hojas
                header=None,
                engine='calamine'
            )
            usar_pandas = True
            st.success(f"✅ Archivo cargado correctamente ({len(hojas_pandas)} hojas)")
            st.info(f"📋 Hojas encontradas: {', '.join(str(k) for k in hojas_pandas.keys())}")
        except Exception as e1:
            error_messages.append(f"calamine: {str(e1)[:80]}")
            
            # INTENTO 2: openpyxl estándar
            try:
                wb = load_workbook(BytesIO(file_content), data_only=True)
                st.success("✅ Archivo cargado con openpyxl")
            except Exception as e2:
                error_messages.append(f"openpyxl (data_only=True): {str(e2)[:80]}")
                
                # INTENTO 3: openpyxl sin data_only
                try:
                    wb = load_workbook(BytesIO(file_content), data_only=False)
                    st.warning("⚠️ Cargado en modo alternativo (openpyxl)")
                except Exception as e3:
                    error_messages.append(f"openpyxl (data_only=False): {str(e3)[:80]}")
                    
                    # INTENTO 4: pandas con openpyxl engine
                    try:
                        hojas_pandas = pd.read_excel(
                            BytesIO(file_content),
                            sheet_name=None,
                            header=None,
                            engine='openpyxl'
                        )
                        usar_pandas = True
                        st.success(f"✅ Archivo cargado con pandas/openpyxl ({len(hojas_pandas)} hojas)")
                    except Exception as e4:
                        error_messages.append(f"pandas/openpyxl: {str(e4)[:80]}")
                        
                        # INTENTO 5: xlrd (para .xls antiguos)
                        try:
                            hojas_pandas = pd.read_excel(
                                BytesIO(file_content),
                                sheet_name=None,
                                header=None,
                                engine='xlrd'
                            )
                            usar_pandas = True
                            st.success(f"✅ Archivo cargado con xlrd ({len(hojas_pandas)} hojas)")
                        except Exception as e5:
                            error_messages.append(f"xlrd: {str(e5)[:80]}")
                            
                            # Si todo falla
                            st.error("❌ No se pudo cargar el archivo")
                            st.markdown("**Errores encontrados:**")
                            for i, msg in enumerate(error_messages, 1):
                                st.text(f"{i}. {msg}")
                            
                            st.markdown("""
                            ### 🔧 Solución recomendada:
                            1. Abre el archivo en **Microsoft Excel**
                            2. Ve a **Archivo → Guardar como**
                            3. Selecciona **Libro de Excel (.xlsx)**
                            4. Guarda con un **nuevo nombre**
                            5. Sube el archivo nuevo
                            """)
                            return None
        
        if not usar_pandas and wb is None:
            st.error("❌ Error inesperado al cargar el archivo")
            return None
            
    except Exception as e:
        st.error(f"❌ Error crítico: {e}")
        traceback.print_exc()
        return None
    
    # Definir funciones auxiliares según el método de carga usado
    if usar_pandas:
        # Usar pandas DataFrames para acceder a los datos
        def get_cell_value(sheet_name, row, col):
            """Obtiene valor de celda usando pandas (row y col son 1-based como Excel)"""
            try:
                # Buscar la hoja (puede ser string o int)
                hoja_df = None
                if sheet_name in hojas_pandas:
                    hoja_df = hojas_pandas[sheet_name]
                else:
                    # Intentar convertir a int si es string numérico
                    try:
                        sheet_int = int(sheet_name)
                        if sheet_int in hojas_pandas:
                            hoja_df = hojas_pandas[sheet_int]
                    except:
                        pass
                
                if hoja_df is None:
                    return ""
                
                # Convertir a 0-based index
                r = row - 1
                c = col - 1
                if r < 0 or c < 0 or r >= len(hoja_df) or c >= len(hoja_df.columns):
                    return ""
                valor = hoja_df.iloc[r, c]
                if pd.isna(valor):
                    return ""
                valor_str = str(valor).strip()
                return valor_str if valor_str and valor_str != "0" and valor_str.lower() != "none" else ""
            except:
                return ""
        
        def sheet_exists(sheet_name):
            if sheet_name in hojas_pandas:
                return True
            try:
                sheet_int = int(sheet_name)
                return sheet_int in hojas_pandas
            except:
                return False
    else:
        # Usar openpyxl workbook para acceder a los datos
        def get_cell_value(sheet_name, row, col):
            """Obtiene valor de celda usando openpyxl (row y col son 1-based)"""
            try:
                if sheet_name not in wb.sheetnames:
                    return ""
                hoja = wb[sheet_name]
                valor = hoja.cell(row=row, column=col).value
                if valor is None:
                    return ""
                valor_str = str(valor).strip()
                if valor_str.startswith('='):  # Es una fórmula sin valor
                    return ""
                return valor_str if valor_str != "0" and valor_str.lower() != "none" else ""
            except:
                return ""
        
        def sheet_exists(sheet_name):
            return sheet_name in wb.sheetnames

    resultado = {
        "informacion_general": {
            "razon_social": "",
            "rut_empresa": "",
            "actividad_economica": "",
            "nombre_centro_trabajo": "",
            "direccion_ct": "",
            "comuna_ct": "",
            "total_trabajadores_hombres": 0,
            "total_trabajadores_mujeres": 0,
            "total_trabajadores": 0
        },
        "puestos_detalle": [],
        "resumen_factores": {},
        "alertas_validacion": {
            "identificacion_inicial": [],
            "evaluaciones_pendientes": []
        }
    }
    
    # ===== VALIDACIÓN HOJA 3: IDENTIFICACIÓN INICIAL =====
    try:
        alertas_id_inicial = validar_identificacion_inicial(get_cell_value, sheet_exists)
        resultado["alertas_validacion"]["identificacion_inicial"] = alertas_id_inicial
    except Exception as e:
        st.warning(f"⚠️ Error al validar identificación inicial: {e}")
    
    # ===== VALIDACIÓN CRUZADA: HOJA 3 vs HOJAS DE FACTORES =====
    try:
        alertas_eval_pendientes = validar_evaluaciones_pendientes(get_cell_value, sheet_exists)
        resultado["alertas_validacion"]["evaluaciones_pendientes"] = alertas_eval_pendientes
    except Exception as e:
        st.warning(f"⚠️ Error al validar evaluaciones pendientes: {e}")

    # ===== PROCESAMIENTO HOJA 1: INFORMACIÓN GENERAL =====
    try:
        if sheet_exists("1"):
            # E=5, L=12, G=7
            resultado["informacion_general"]["razon_social"] = get_cell_value("1", 15, 5)  # E15
            resultado["informacion_general"]["rut_empresa"] = get_cell_value("1", 15, 12)  # L15
            resultado["informacion_general"]["actividad_economica"] = get_cell_value("1", 17, 5)  # E17
            resultado["informacion_general"]["nombre_centro_trabajo"] = get_cell_value("1", 27, 5)  # E27
            resultado["informacion_general"]["direccion_ct"] = get_cell_value("1", 29, 5)  # E29
            resultado["informacion_general"]["comuna_ct"] = get_cell_value("1", 29, 12)  # L29
            
            hombres_str = get_cell_value("1", 31, 7)  # G31
            mujeres_str = get_cell_value("1", 31, 12)  # L31
            try:
                hombres = int(float(hombres_str)) if hombres_str else 0
                mujeres = int(float(mujeres_str)) if mujeres_str else 0
            except:
                hombres = 0
                mujeres = 0
            
            resultado["informacion_general"]["total_trabajadores_hombres"] = hombres
            resultado["informacion_general"]["total_trabajadores_mujeres"] = mujeres
            resultado["informacion_general"]["total_trabajadores"] = hombres + mujeres
        else:
            st.warning("⚠️ No se encontró la Hoja '1'. Información general incompleta.")
    except Exception as e:
        st.error(f"Error procesando Hoja '1': {e}")
        traceback.print_exc()

    # ===== CONFIGURACIÓN DE FACTORES DE RIESGO =====
    agentes_riesgo = [
        "Repetitividad", "Postura", "MMC LDT", "MMC EA", 
        "MMP", "Vibración MB", "Vibración CC"
    ]
    
    # Inicializar resumen de factores
    for agente in agentes_riesgo:
        resultado["resumen_factores"][agente] = {
            "nivel_maximo": "AUSENTE",
            "puestos_criticos": [],
            "tiene_plan_accion": None,  # None = no aplica, True = completo, False = sin plan, 'incompleto'
            "prioridad": 0,  # 0=ausente, 1=aceptable, 2=intermedio, 3=crítico
            # Contadores de casos por nivel
            "casos_critico": 0,
            "casos_intermedio": 0,
            "casos_aceptable": 0,
            "casos_ausente": 0
        }

    # Mapeo de prioridades
    prioridad_map = {
        "AUSENTE": 0,
        "ACEPTABLE": 1,
        "INTERMEDIO": 2,
        "CRÍTICO": 3,
        "NO ACEPTABLE": 3
    }

    # ===== PROCESAMIENTO HOJA 2: PUESTOS DE TRABAJO =====
    mapa_nro_puesto = {}
    try:
        if sheet_exists("2"):
            COL_NRO, COL_AREA, COL_PUESTO, COL_TAREA = 2, 3, 4, 5  # B, C, D, E
            
            for fila_idx in range(13, 114):
                nro_puesto = get_cell_value("2", fila_idx, COL_NRO)
                area = get_cell_value("2", fila_idx, COL_AREA)
                puesto = get_cell_value("2", fila_idx, COL_PUESTO)
                tarea = get_cell_value("2", fila_idx, COL_TAREA)
                
                if not (nro_puesto and area and puesto):
                    continue
                
                puesto_info = {
                    "nro": nro_puesto,
                    "area": area,
                    "puesto": puesto,
                    "tarea": tarea,
                    "niveles_riesgo": {normalize_key(ag): "AUSENTE" for ag in agentes_riesgo}
                }
                
                resultado["puestos_detalle"].append(puesto_info)
                mapa_nro_puesto[nro_puesto] = len(resultado["puestos_detalle"]) - 1
        else:
            st.warning("⚠️ No se encontró la Hoja '2'. No se cargarán puestos de trabajo.")
    except Exception as e:
        st.error(f"Error procesando Hoja '2': {e}")
        traceback.print_exc()

    # ===== CONFIGURACIÓN DE HOJAS DE FACTORES =====
    config_hojas_factores = {
        "4": {
            "agente": "Repetitividad",
            "col_q_idx": 17,
            "col_x_idx": 24,
            "r_filas": (14, 116),
            "columnas_plan": ["Y", "Z", "AB", "AC", "AD"]
        },
        "5": {
            "agente": "Postura",
            "col_q_idx": 31,
            "col_x_idx": 49,
            "r_filas": (17, 116),
            "columnas_plan": ["AX", "AY", "BA", "BB", "BC"]
        },
        "6": {
            "agente": "MMC LDT",
            "col_q_idx": 33,
            "col_x_idx": 56,
            "r_filas": (18, 118),
            "columnas_plan": ["BE", "BF", "BH", "BI", "BJ"]
        },
        "7": {
            "agente": "MMC EA",
            "col_q_idx": 24,
            "col_x_idx": 41,
            "r_filas": (17, 117),
            "columnas_plan": ["AP", "AQ", "AS", "AT", "AU"]
        },
        "8": {
            "agente": "MMP",
            "col_q_idx": 25,
            "col_x_idx": 41,
            "r_filas": (17, 117),
            "columnas_plan": ["AP", "AQ", "AS", "AT", "AU"]
        },
        "9": {
            "agente": "Vibración MB",
            "col_riesgo_directo_idx": 19,
            "r_filas": (16, 116),
            "columnas_plan": ["T", "U", "W", "X", "Y"]
        },
        "10": {
            "agente": "Vibración CC",
            "col_riesgo_directo_idx": 22,
            "r_filas": (16, 116),
            "columnas_plan": ["W", "X", "Z", "AA", "AB"]
        }
    }

    # ===== PROCESAMIENTO DE HOJAS DE FACTORES =====
    for num_hoja, config in config_hojas_factores.items():
        try:
            if not sheet_exists(num_hoja):
                st.warning(f"⚠️ No se encontró la Hoja '{num_hoja}' ({config['agente']})")
                continue
                
            agente = config["agente"]
            agente_norm = normalize_key(agente)
            COL_NRO, COL_AREA, COL_PUESTO = 2, 3, 4  # B, C, D
            
            puestos_criticos_con_plan = []
            
            for fila_idx in range(config["r_filas"][0], config["r_filas"][1]):
                nro_puesto = get_cell_value(num_hoja, fila_idx, COL_NRO)
                area = get_cell_value(num_hoja, fila_idx, COL_AREA)
                puesto = get_cell_value(num_hoja, fila_idx, COL_PUESTO)
                
                if not (nro_puesto and area and puesto):
                    continue
                
                # Determinar nivel de riesgo
                risk_level = "AUSENTE"
                
                if "col_riesgo_directo_idx" in config:
                    # Vibraciones: Aceptable o No Aceptable
                    valor_str = get_cell_value(num_hoja, fila_idx, config["col_riesgo_directo_idx"]).lower()
                    if valor_str == "aceptable":
                        risk_level = "ACEPTABLE"
                    elif valor_str == "no aceptable":
                        risk_level = "CRÍTICO"
                else:
                    # Otros factores: Aceptable, Intermedio, Crítico
                    valor_q_str = get_cell_value(num_hoja, fila_idx, config["col_q_idx"]).lower()
                    
                    if valor_q_str == "no aceptable":
                        valor_x_str = get_cell_value(num_hoja, fila_idx, config["col_x_idx"]).lower()
                        if "no crítico" in valor_x_str or "intermedio" in valor_x_str:
                            risk_level = "INTERMEDIO"
                        elif "crítico" in valor_x_str:
                            risk_level = "CRÍTICO"
                    elif valor_q_str == "aceptable":
                        risk_level = "ACEPTABLE"
                
                # Actualizar nivel en puesto
                if nro_puesto in mapa_nro_puesto:
                    idx = mapa_nro_puesto[nro_puesto]
                    resultado["puestos_detalle"][idx]["niveles_riesgo"][agente_norm] = risk_level
                
                # Incrementar contador de casos por nivel
                if risk_level == "CRÍTICO":
                    resultado["resumen_factores"][agente]["casos_critico"] += 1
                elif risk_level == "INTERMEDIO":
                    resultado["resumen_factores"][agente]["casos_intermedio"] += 1
                elif risk_level == "ACEPTABLE":
                    resultado["resumen_factores"][agente]["casos_aceptable"] += 1
                else:
                    resultado["resumen_factores"][agente]["casos_ausente"] += 1
                
                # Actualizar nivel máximo del factor
                prioridad_actual = prioridad_map.get(risk_level, 0)
                prioridad_maxima = resultado["resumen_factores"][agente]["prioridad"]
                
                if prioridad_actual > prioridad_maxima:
                    resultado["resumen_factores"][agente]["nivel_maximo"] = risk_level
                    resultado["resumen_factores"][agente]["prioridad"] = prioridad_actual
                    resultado["resumen_factores"][agente]["puestos_criticos"] = []
                
                # Si es crítico, agregar a lista y validar plan
                if risk_level == "CRÍTICO":
                    info_puesto = f"Puesto {nro_puesto}: {area} - {puesto}"
                    resultado["resumen_factores"][agente]["puestos_criticos"].append(info_puesto)
                    
                    # Validar plan de acción
                    estado_plan = validar_plan_accion_general(get_cell_value, num_hoja, fila_idx, config["columnas_plan"])
                    puestos_criticos_con_plan.append({
                        "puesto": info_puesto,
                        "estado_plan": estado_plan
                    })
            
            # Determinar estado general del plan para este factor
            if resultado["resumen_factores"][agente]["nivel_maximo"] == "CRÍTICO":
                if puestos_criticos_con_plan:
                    # Verificar si todos los puestos críticos tienen plan completo
                    todos_completos = all(p["estado_plan"] == "completo" for p in puestos_criticos_con_plan)
                    alguno_sin_plan = any(p["estado_plan"] == "sin_plan" for p in puestos_criticos_con_plan)
                    
                    if todos_completos:
                        resultado["resumen_factores"][agente]["tiene_plan_accion"] = True
                    elif alguno_sin_plan:
                        resultado["resumen_factores"][agente]["tiene_plan_accion"] = False
                    else:
                        resultado["resumen_factores"][agente]["tiene_plan_accion"] = "incompleto"
                    
                    resultado["resumen_factores"][agente]["detalle_planes"] = puestos_criticos_con_plan
                        
        except Exception as e:
            st.error(f"Error procesando Hoja '{num_hoja}': {e}")
            traceback.print_exc()

    return resultado

# ===== INTERFAZ STREAMLIT =====
st.set_page_config(
    page_title="Resumen Matriz TMERT",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.title("📊 Resumen Ejecutivo - Matriz TMERT")
st.markdown("**Desarrollado por Mauricio Reyes González**")
st.markdown("---")

# Cargar archivo
uploaded_file = st.file_uploader(
    "📤 Cargar Matriz TMERT (Excel versión 7 de ACHS)",
    type=["xlsx"],
    help="Selecciona el archivo Excel con la matriz TMERT completa"
)

if uploaded_file:
    with st.spinner("⚙️ Procesando archivo Excel..."):
        datos = procesar_excel_resumen(uploaded_file)
    
    if datos:
        st.success("✅ Archivo procesado correctamente")
        
        # ===== RESUMEN EJECUTIVO SIMPLIFICADO =====
        st.markdown("## 📊 Resumen Ejecutivo")
        
        info = datos["informacion_general"]
        resumen = datos["resumen_factores"]
        alertas_id_inicial = datos.get("alertas_validacion", {}).get("identificacion_inicial", [])
        alertas_eval_pendientes = datos.get("alertas_validacion", {}).get("evaluaciones_pendientes", [])
        
        # Calcular estadísticas de CASOS (no factores)
        total_alertas = len(alertas_id_inicial) + len(alertas_eval_pendientes)
        
        # Contar total de casos por nivel (sumando todos los factores)
        total_casos_criticos = sum(resumen[f]["casos_critico"] for f in resumen)
        total_casos_intermedios = sum(resumen[f]["casos_intermedio"] for f in resumen)
        total_casos_aceptables = sum(resumen[f]["casos_aceptable"] for f in resumen)
        
        # Factores con nivel máximo crítico (para planes de acción)
        factores_criticos = [f for f, data in resumen.items() if data["nivel_maximo"] == "CRÍTICO"]
        factores_criticos_con_plan = [f for f in factores_criticos if resumen[f]["tiene_plan_accion"] == True]
        factores_criticos_sin_plan = [f for f in factores_criticos if resumen[f]["tiene_plan_accion"] == False]
        
        # Tarjeta de resumen compacta
        col_emp, col_estado = st.columns([1, 2])
        
        with col_emp:
            st.markdown("### 🏢 Empresa")
            st.markdown(f"**{info['razon_social']}**")
            st.markdown(f"RUT: {info['rut_empresa']}")
            st.markdown(f"CT: {info['nombre_centro_trabajo']}")
            st.markdown(f"Trabajadores: **{info['total_trabajadores']}**")
        
        with col_estado:
            st.markdown("### 📋 Estado de la Revisión")
            
            # Indicadores en una fila
            col_a, col_b, col_c, col_d = st.columns(4)
            
            with col_a:
                color_alertas = "🔴" if total_alertas > 0 else "🟢"
                st.metric(f"{color_alertas} Alertas", total_alertas)
            
            with col_b:
                st.metric("🔴 Casos Críticos", total_casos_criticos)
            
            with col_c:
                st.metric("🟡 Casos Intermedios", total_casos_intermedios)
            
            with col_d:
                if len(factores_criticos) > 0:
                    cumplimiento = int((len(factores_criticos_con_plan) / len(factores_criticos)) * 100)
                    st.metric("📝 Planes Acción", f"{cumplimiento}%")
                else:
                    st.metric("📝 Planes Acción", "N/A")
        
        # Resumen de factores de riesgo con conteos
        st.markdown("### 🎯 Factores de Riesgo")
        
        factores_orden = ["Repetitividad", "Postura", "MMC LDT", "MMC EA", "MMP", "Vibración MB", "Vibración CC"]
        
        # Crear tabla de resumen por factor
        tabla_factores = []
        for factor in factores_orden:
            data = resumen[factor]
            tabla_factores.append({
                "Factor": factor,
                "🔴 Crítico": data["casos_critico"],
                "🟡 Intermedio": data["casos_intermedio"],
                "🟢 Aceptable": data["casos_aceptable"],
                "Nivel Máximo": data["nivel_maximo"]
            })
        
        df_factores = pd.DataFrame(tabla_factores)
        
        # Función para colorear según nivel máximo
        def colorear_nivel_maximo(row):
            nivel = row["Nivel Máximo"]
            if nivel == "CRÍTICO":
                return ['background-color: #ffcccc'] * len(row)
            elif nivel == "INTERMEDIO":
                return ['background-color: #fff3cd'] * len(row)
            elif nivel == "ACEPTABLE":
                return ['background-color: #d4edda'] * len(row)
            else:
                return ['background-color: #f8f9fa'] * len(row)
        
        df_styled = df_factores.style.apply(colorear_nivel_maximo, axis=1)
        st.dataframe(df_styled, use_container_width=True, hide_index=True)
        
        st.markdown("---")
        
        # ===== SECCIÓN ALERTAS DE VALIDACIÓN (DETALLE) =====
        tiene_alertas = alertas_id_inicial or alertas_eval_pendientes
        
        if tiene_alertas:
            st.markdown("## ⚠️ Alertas de Validación")
            
            # Alerta 1: Identificación Inicial Incompleta
            if alertas_id_inicial:
                with st.expander(f"🔍 Identificación Inicial Incompleta: {len(alertas_id_inicial)} caso(s)", expanded=False):
                    st.warning(f"Se encontraron **{len(alertas_id_inicial)} caso(s)** con identificación inicial incompleta en la Hoja 3.")
                    st.markdown("**Casos afectados:**")
                    
                    casos_por_mostrar = alertas_id_inicial[:50]
                    
                    df_alertas = pd.DataFrame([
                        {
                            "Caso": alerta["caso"],
                            "Fila": alerta["fila"],
                            "Columnas Faltantes": ", ".join(alerta["columnas_faltantes"])
                        }
                        for alerta in casos_por_mostrar
                    ])
                    
                    st.dataframe(df_alertas, use_container_width=True, hide_index=True)
                    
                    if len(alertas_id_inicial) > 50:
                        st.info(f"ℹ️ Mostrando los primeros 50 casos de {len(alertas_id_inicial)} totales.")
            
            # Alerta 2: Evaluaciones Pendientes
            if alertas_eval_pendientes:
                with st.expander(f"📋 Evaluaciones Pendientes: {len(alertas_eval_pendientes)} caso(s)", expanded=False):
                    st.warning(f"Se encontraron **{len(alertas_eval_pendientes)}** evaluaciones pendientes (marcados SI en Hoja 3 pero sin evaluación).")
                    st.markdown("**Evaluaciones faltantes:**")
                    
                    casos_por_mostrar = alertas_eval_pendientes[:50]
                    
                    df_eval_pendientes = pd.DataFrame([
                        {
                            "Caso": alerta["caso"],
                            "Factor de Riesgo": alerta["factor"],
                            "Hoja Faltante": alerta["hoja"]
                        }
                        for alerta in casos_por_mostrar
                    ])
                    
                    st.dataframe(df_eval_pendientes, use_container_width=True, hide_index=True)
                    
                    if len(alertas_eval_pendientes) > 50:
                        st.info(f"ℹ️ Mostrando los primeros 50 casos de {len(alertas_eval_pendientes)} totales.")
            
            st.markdown("---")
        
        # ===== SECCIÓN DETALLE: TABLA DE RIESGOS =====
        st.markdown("## 📋 Detalle de Factores de Riesgo")
        
        # Crear datos para la tabla
        tabla_data = []
        for factor in ["Repetitividad", "Postura", "MMC LDT", "MMC EA", "MMP", "Vibración MB", "Vibración CC"]:
            data = resumen[factor]
            nivel = data["nivel_maximo"]
            
            # Emoji según nivel
            if nivel == "CRÍTICO":
                nivel_display = "🔴 CRÍTICO"
            elif nivel == "INTERMEDIO":
                nivel_display = "🟡 INTERMEDIO"
            elif nivel == "ACEPTABLE":
                nivel_display = "🟢 ACEPTABLE"
            else:
                nivel_display = "⚪ AUSENTE"
            
            # Estado del plan
            if data["tiene_plan_accion"] is None:
                plan_display = "➖ No aplica"
            elif data["tiene_plan_accion"] == True:
                plan_display = "✅ Completo"
            elif data["tiene_plan_accion"] == False:
                plan_display = "❌ Sin registrar"
            else:
                plan_display = "⚠️ Incompleto"
            
            num_puestos = len(data["puestos_criticos"])
            
            tabla_data.append({
                "Factor de Riesgo": factor,
                "Nivel Máximo": nivel_display,
                "Plan de Acción": plan_display,
                "N° Puestos Críticos": num_puestos if num_puestos > 0 else "-"
            })
        
        # Mostrar como DataFrame con estilo
        df_resumen = pd.DataFrame(tabla_data)
        
        # Función para colorear las filas según nivel de riesgo
        def colorear_fila(row):
            nivel = row["Nivel Máximo"]
            if "CRÍTICO" in nivel:
                return ['background-color: #ffcccc'] * len(row)
            elif "INTERMEDIO" in nivel:
                return ['background-color: #fff3cd'] * len(row)
            elif "ACEPTABLE" in nivel:
                return ['background-color: #d4edda'] * len(row)
            else:
                return ['background-color: #f8f9fa'] * len(row)
        
        # Aplicar estilo y mostrar
        df_styled = df_resumen.style.apply(colorear_fila, axis=1)
        st.dataframe(
            df_styled,
            use_container_width=True,
            hide_index=True,
            height=300
        )
        
        # ===== ALERTAS =====
        if factores_criticos_sin_plan:
            st.markdown("### ⚠️ Alertas Críticas")
            for factor in factores_criticos_sin_plan:
                st.error(f"**{factor}** tiene riesgo CRÍTICO sin plan de acción registrado")
                # Mostrar puestos afectados
                puestos = resumen[factor]["puestos_criticos"]
                with st.expander(f"Ver {len(puestos)} puesto(s) afectado(s)"):
                    for puesto in puestos:
                        st.write(f"• {puesto}")
        
        if len(factores_criticos) > 0:
            porcentaje_cumplimiento = (len(factores_criticos_con_plan) / len(factores_criticos)) * 100
            st.markdown(f"### 📊 Cumplimiento de Planes de Acción: **{porcentaje_cumplimiento:.0f}%**")
            st.progress(porcentaje_cumplimiento / 100)
        
        st.markdown("---")
        
        # ===== SECCIÓN: INFORMACIÓN DETALLADA DE LA EMPRESA =====
        with st.expander("🏢 Ver Información Completa de la Empresa"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 📋 Antecedentes de la Empresa")
                st.markdown(f"**Razón Social:** {info['razon_social']}")
                st.markdown(f"**RUT:** {info['rut_empresa']}")
                st.markdown(f"**Actividad Económica:** {info['actividad_economica']}")
            
            with col2:
                st.markdown("### 📍 Centro de Trabajo")
                st.markdown(f"**Nombre:** {info['nombre_centro_trabajo']}")
                st.markdown(f"**Dirección:** {info['direccion_ct']}")
                st.markdown(f"**Comuna:** {info['comuna_ct']}")
            
            st.markdown("### 👥 Dotación")
            col_t1, col_t2, col_t3 = st.columns(3)
            with col_t1:
                st.metric("Total Trabajadores", info['total_trabajadores'])
            with col_t2:
                st.metric("Hombres", info['total_trabajadores_hombres'])
            with col_t3:
                st.metric("Mujeres", info['total_trabajadores_mujeres'])
        
        # ===== SECCIÓN: DETALLE POR PUESTO (OPCIONAL) =====
        with st.expander("📑 Ver Detalle por Puesto de Trabajo"):
            st.markdown("### Detalle de Todos los Puestos Evaluados")
            
            for puesto in datos["puestos_detalle"]:
                st.markdown(f"#### Puesto {puesto['nro']}: {puesto['puesto']}")
                st.markdown(f"**Área:** {puesto['area']}")
                st.markdown(f"**Tarea:** {puesto['tarea']}")
                
                # Mostrar niveles de riesgo
                col_d1, col_d2, col_d3, col_d4 = st.columns(4)
                riesgos = puesto['niveles_riesgo']
                
                factores = list(riesgos.keys())
                for i, factor_norm in enumerate(factores):
                    nivel = riesgos[factor_norm]
                    
                    # Encontrar nombre original del factor
                    factor_original = ""
                    for f in ["Repetitividad", "Postura", "MMC LDT", "MMC EA", "MMP", "Vibración MB", "Vibración CC"]:
                        if normalize_key(f) == factor_norm:
                            factor_original = f
                            break
                    
                    if nivel == "CRÍTICO":
                        emoji = "🔴"
                    elif nivel == "INTERMEDIO":
                        emoji = "🟡"
                    elif nivel == "ACEPTABLE":
                        emoji = "🟢"
                    else:
                        emoji = "⚪"
                    
                    col = [col_d1, col_d2, col_d3, col_d4][i % 4]
                    with col:
                        st.markdown(f"**{factor_original}**  \n{emoji} {nivel}")
                
                st.markdown("---")
        
        # ===== PIE DE PÁGINA =====
        st.markdown("---")
        st.markdown(f"**Fecha de generación:** {datetime.now().strftime('%d-%m-%Y %H:%M')}")
        st.markdown(f"**Archivo procesado:** {uploaded_file.name}")

else:
    st.info("👆 Por favor, carga un archivo Excel para comenzar el análisis")
    
    # Instrucciones
    with st.expander("ℹ️ Instrucciones de Uso"):
        st.markdown("""
        ### Cómo usar esta aplicación:
        
        1. **Cargar archivo Excel**: Utiliza el botón de carga para seleccionar tu matriz TMERT (versión 7 de ACHS)
        
        2. **Revisión automática**: La aplicación procesará automáticamente:
           - Información general de la empresa
           - Validación de identificación inicial (Hoja 3)
           - Validación cruzada entre Hoja 3 y hojas de evaluación
           - Niveles de riesgo por cada factor
           - Validación de planes de acción para riesgos críticos
        
        3. **Resultados**: Verás:
           - **Resumen Ejecutivo**: Datos de empresa y estado general
           - **Alertas de Validación**: Problemas encontrados en la matriz
           - **Tabla de Riesgos**: Nivel máximo por factor con colores
           - **Cumplimiento**: Porcentaje de planes de acción registrados
           - **Detalles**: Información expandible por puesto
        
        ### Validaciones que realiza:
        - ✅ Identificación inicial completa (SI/NO en columnas E-K de Hoja 3)
        - ✅ Evaluaciones pendientes (SI en Hoja 3 pero sin evaluación en hoja correspondiente)
        - ✅ Planes de acción para riesgos críticos
        
        ### Factores de Riesgo Evaluados:
        - Repetitividad (Hoja 4)
        - Postura (Hoja 5)
        - MMC LDT - Levantamiento y Descenso (Hoja 6)
        - MMC EA - Empuje y Arrastre (Hoja 7)
        - MMP - Manejo Manual de Pacientes (Hoja 8)
        - Vibración MB - Mano-Brazo (Hoja 9)
        - Vibración CC - Cuerpo Completo (Hoja 10)
        
        ### Niveles de Riesgo:
        - 🔴 **CRÍTICO**: Requiere acción inmediata y plan de acción
        - 🟡 **INTERMEDIO**: Requiere seguimiento
        - 🟢 **ACEPTABLE**: Dentro de límites permitidos
        - ⚪ **AUSENTE**: No se detectó exposición
        """)

